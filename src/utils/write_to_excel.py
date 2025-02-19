import json
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def write_to_excel(json_data, output_file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers =  json_data[0].keys()

    # Add headers
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=json_data[0].get(header,{}).get('title'))

    # Add data rows
    for row_num, item in enumerate(json_data, 2):
        for col_num, header in enumerate(headers, 1):
            data = item[header]
            cell = sheet.cell(row=row_num, column=col_num)
            if data.get("href"):
                # Set cell as clickable hyperlink
                cell.value = data["value"]
                cell.hyperlink = data["href"]
                cell.font = Font(color="0000FF", underline="single")
            else:
                cell.value = data["value"]
                color = data.get("textColor", "FF000000")
                cell.font = Font(color=color)

    # Adjust column widths
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        # Add a little extra space
        sheet.column_dimensions[col_letter].width = max_length + 2

    workbook.save(output_file_path)

if __name__ == "__main__":
    with open('example.json', 'r', encoding='utf-8') as f:
        json_data = json.load(f)
    write_to_excel(json_data, 'output.xlsx')
